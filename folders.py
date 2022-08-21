import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
from openpyxl import load_workbook


class LIVEChallengeFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'Images', imgpath[item][0][0]), labels[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target / 100

    def __len__(self):
        length = len(self.samples)
        return length

class BIDFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class Koniq_10kFolder(data.Dataset):


    def __init__(self, root, index, transform, patch_num):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, '512x384', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class SPAQFolder(data.Dataset):


    def __init__(self, root, index, transform, patch_num):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'MOS and Image attribute scores.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['Image name'])
                mos = np.array(float(row['MOS'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'TestImage', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class FLIVEFolder(data.Dataset):


    def __init__(self, root, index, transform, transform_s, patch_num):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'labels_image.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['name'])
                mos = np.array(float(row['mos'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'database', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform
        self.transform_s = transform_s

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        h, w = sample.size
        if min(h, w) > 224:
            sample = self.transform(sample)
        else:
            sample = self.transform_s(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class LSRQFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num, istrain):
        imgnamea = []
        imgnameb = []
        label_all = []
        csv_file = os.path.join(root, 'pseudo-labels.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgnamea.append(row['imagenamea'])
                imgnameb.append(row['imagenameb'])
                label = np.array(float(row['label'])).astype(np.float32)
                label_all.append(label)

        samplea = []
        sampleb = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                samplea.append((os.path.join(root, 'Images', imgnamea[item]), label_all[item]))
                sampleb.append((os.path.join(root, 'Images', imgnameb[item]), label_all[item]))

        self.samplea = samplea
        self.sampleb = sampleb
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        patha, target = self.samplea[index]
        pathb, target = self.sampleb[index]
        samplea = pil_loader(patha)
        sampleb = pil_loader(pathb)
        samplea = self.transform(samplea)
        sampleb = self.transform(sampleb)
        return samplea, sampleb, target

    def __len__(self):
        length = len(self.samplea)
        return length


def pil_loader(path):
    with open(path, 'rb') as f:
        img = Image.open(f)
        return img.convert('RGB')